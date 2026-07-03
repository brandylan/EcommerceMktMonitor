"""
nint.com 爬虫 - 阿里天猫数据采集（增量更新）
依赖：pip install selenium openpyxl
"""

import pandas as pd
import math
import os, re, time, traceback, shutil, json, subprocess, sys
from pathlib import Path
from datetime import datetime, timedelta
from html import unescape
from urllib.parse import parse_qs, urlparse
from collections import defaultdict
from typing import Dict, List, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

# ─── 配置 ────────────────────────────────────────────────
EMAIL    = os.environ.get("NINT_EMAIL",    "Ding.YUAN@cn.bosch.com")
PASSWORD = os.environ.get("NINT_PASSWORD", "Bosch2025")
LOGIN_URL = "https://account.nint.com/"
WAIT = 15
HEADLESS = os.environ.get("NINT_HEADLESS", "1").strip().lower() in {"1", "true", "yes", "y"}
ASSUME_PROXY_READY = os.environ.get("NINT_ASSUME_PROXY_READY", "0").strip().lower() in {"1", "true", "yes", "y"}
CHROME_USER_DATA_DIR = os.environ.get("NINT_CHROME_USER_DATA_DIR", "").strip()
CHROME_PROFILE_DIR = os.environ.get("NINT_CHROME_PROFILE_DIR", "").strip()
ALLOW_HEADED_FALLBACK = os.environ.get("NINT_ALLOW_HEADED_FALLBACK", "0").strip().lower() in {"1", "true", "yes", "y"}

EXCLUDE_KEYWORDS_TMALL = ["自定义价格段分析", "新建自定义属性", "类目变动查询", "温馨提示"]
EXCLUDE_KEYWORDS_JD    = ["自定义价格段分析", "新建自定义属性", "类目变动查询", "温馨提示"]
JD_EXACT_EXCLUDE = ["机油滤", "MBL (收藏)", "MBL"]   # 精确匹配排除
LEAF_KEYWORDS    = ["热销品牌"]

PRODUCT_NAME_MAP = {
    "雨刮器": "雨刷", "空调滤芯": "空调滤清器",
    "机油滤芯": "机油滤清器", "空气滤芯": "空气滤清器",
    "汽油滤芯": "燃油滤清器", "汽车电瓶/蓄电池": "蓄电池", "蓄电池": "蓄电池",
}

# ── 固定品类 URL（跳过侧边栏递归，直接采集）──────────────
TMALL_URLS = {
    "汽机油":    "https://art.nint.com/stat-ali-new?cid=merge_260212&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "空气滤清器": "https://art.nint.com/stat-ali-new?cid=merge_50016288&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "燃油滤清器": "https://art.nint.com/stat-ali-new?cid=merge_50016289&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "机油滤清器": "https://art.nint.com/stat-ali-new?cid=merge_50015525&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "空调滤清器": "https://art.nint.com/stat-ali-new?cid=merge_50010627&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "火花塞":    "https://art.nint.com/stat-ali-new?cid=merge_50012859&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "点火线圈":   "https://art.nint.com/stat-ali-new?cid=merge_50014560&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "喷油嘴":    "https://art.nint.com/stat-ali-new?cid=merge_50012865&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "雨刷":      "https://art.nint.com/stat-ali-new?cid=merge_50012888&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "蓄电池":    "https://art.nint.com/stat-ali-new?cid=merge_50014563&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
    "刹车片":    "https://art.nint.com/stat-ali-new?cid=merge_50014566&site=ali&rcid=zdy_cid&zdy_cid=zdy_14394#fold_line",
}

JD_URLS = {
    "雨刷":      "https://art.nint.com/stat-new-v2?cid=merge_6766&site=jd&zdy_cid=zdy_4621#fold_line",
    "火花塞":    "https://art.nint.com/stat-new-v2?cid=merge_6767&site=jd&zdy_cid=zdy_4621#fold_line",
    "蓄电池":    "https://art.nint.com/stat-new-v2?cid=merge_9971&site=jd&zdy_cid=zdy_4621#fold_line",
    "汽机油":    "https://art.nint.com/stat-new-v2?cid=merge_11849&site=jd&zdy_cid=zdy_4621#fold_line",
    "机油滤清器": "https://art.nint.com/stat-new-v2?cid=merge_11852&site=jd&zdy_cid=zdy_4621#fold_line",
    "刹车片":    "https://art.nint.com/stat-new-v2?cid=merge_11859&site=jd&zdy_cid=zdy_4621#fold_line",
    "空调滤清器": "https://art.nint.com/stat-new-v2?cid=merge_14888&site=jd&zdy_cid=zdy_4621#fold_line",
    "空气滤清器": "https://art.nint.com/stat-new-v2?cid=merge_14889&site=jd&zdy_cid=zdy_4621#fold_line",
    "燃油滤清器": "https://art.nint.com/stat-new-v2?cid=merge_14890&site=jd&zdy_cid=zdy_4621#fold_line",
    "点火线圈":   "https://art.nint.com/stat-new-v2?cid=merge_16877&site=jd&zdy_cid=zdy_4621#fold_line",
    "氧传感器":   "https://art.nint.com/stat-new-v2?cid=merge_16879&site=jd&zdy_cid=zdy_4621#fold_line",
}
# MktVolumeSummary 18列
HDR_INDUSTRY = [
    "渠道", "时间", "产品",
    "行业类目销售额", "销售额同比",
    "行业类目销量",  "销量同比",
    "品牌总数量", " TOP10品牌类目份额占比 (销售额)",

    "博世_销售额", "博世_销售额同比",
    "博世_销量",  "博世_销量同比",
    "博世_市占比 (销售额)",
    "null",                   # 第15列
    "往年YTD", "往年销售额", "往年销量"
]
# MktShare 5列
HDR_MKT   = ["渠道", "产品", "品牌", "时间", "MktShare"]
HDR_BRAND = ["渠道", "产品", "品牌", "时间", "销量", "销量同比", "销售额", "销售额同比", "均价"]
# ─────────────────────────────────────────────────────────

TARGET_MONTH    = ""
CHANNEL         = "天猫"   # 当前采集渠道：天猫 or 京东
new_industry:   List[Dict] = []
new_brand:      List[Dict] = []
new_mkt:        List[Dict] = []
prev_year_data: Dict       = {}
leaf_results:   List       = []
visited_urls:   Set        = set()


# ══════════════════════════════════════════════════════════
# 基础工具
# ══════════════════════════════════════════════════════════
def init_driver():
    return create_driver(HEADLESS)


def create_driver(headless=False):
    opt = webdriver.ChromeOptions()
    opt.add_argument("--no-sandbox")
    opt.add_argument("--disable-dev-shm-usage")
    opt.add_argument("--window-size=1920,1080")
    if headless:
        opt.add_argument("--headless=new")
        opt.add_argument("--disable-gpu")
    if CHROME_USER_DATA_DIR:
        opt.add_argument(f"--user-data-dir={CHROME_USER_DATA_DIR}")
    if CHROME_PROFILE_DIR:
        opt.add_argument(f"--profile-directory={CHROME_PROFILE_DIR}")
    opt.add_experimental_option("excludeSwitches", ["enable-automation"])
    opt.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(service=Service(r"C:\chromedriver-win64\chromedriver.exe"), options=opt)
    driver.implicitly_wait(2)
    return driver


def init_driver_with_fallback():
    try:
        return create_driver(HEADLESS), HEADLESS
    except Exception as exc:
        if not HEADLESS:
            raise
        if not ALLOW_HEADED_FALLBACK:
            raise RuntimeError(f"无头模式启动失败，且已禁用有头回退：{exc}") from exc
        print(f"⚠ 无头模式启动失败，回退到有头模式：{exc}")
        return create_driver(False), False

def wait_click(driver, xpath, timeout=WAIT):
    el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    el.click(); return el

def wait_input(driver, xpath, text, timeout=WAIT):
    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))
    el.clear(); el.send_keys(text); return el

def map_product(path):
    # path 现在直接就是产品名（如「汽机油」「机油滤清器」），无需再分割
    for k, v in PRODUCT_NAME_MAP.items():
        if k in path: return v
    return path

def parse_num(s):
    try: return float(str(s).replace(",", "").replace("，", ""))
    except: return 0.0

def calc_avg(s, q):
    try:
        sv, qv = parse_num(s), parse_num(q)
        return round(sv / qv, 2) if qv > 0 else ""
    except: return ""

def get_ytd_label(m): return f"{m[:4]}YTD{m[4:]}"

def wait_loading(driver):
    for xpath in ["//*[contains(@class,'loading')]", "//*[contains(@class,'el-loading')]"]:
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath)))
            WebDriverWait(driver, 30).until(EC.invisibility_of_element_located((By.XPATH, xpath)))
            break
        except TimeoutException: continue
    time.sleep(2)


def extract_balanced(text, anchor, open_char, close_char):
    anchor_idx = text.find(anchor)
    if anchor_idx < 0:
        raise ValueError(f"未找到 {anchor}")
    start_idx = text.find(open_char, anchor_idx + len(anchor))
    if start_idx < 0:
        raise ValueError(f"未找到 {anchor} 后的 {open_char}")

    depth = 0
    quote = None
    escape = False
    for idx in range(start_idx, len(text)):
        ch = text[idx]
        if quote:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == quote:
                quote = None
            continue

        if ch in ('"', "'"):
            quote = ch
            continue
        if ch == open_char:
            depth += 1
            continue
        if ch == close_char:
            depth -= 1
            if depth == 0:
                return text[start_idx:idx + 1]

    raise ValueError(f"未能完整解析 {anchor}")


def clean_html_text(text):
    text = re.sub(r"<[^>]+>", "", text)
    text = unescape(text)
    text = text.replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_summary_metrics(html):
    match = re.search(
        r'<div class="form-set describe">.*?<table>.*?<tbody>(.*?)</tbody>.*?</table>',
        html,
        re.S,
    )
    if not match:
        return {}

    tbody = match.group(1)
    rows = re.findall(r"<tr class=\"hot_font\">(.*?)</tr>", tbody, re.S)
    if len(rows) < 2:
        return {}

    sales_vals = [clean_html_text(v) for v in re.findall(r"<td[^>]*>(.*?)</td>", rows[0], re.S)]
    qty_vals = [clean_html_text(v) for v in re.findall(r"<td[^>]*>(.*?)</td>", rows[1], re.S)]
    return {
        "行业类目销售额": sales_vals[0] if len(sales_vals) > 0 else "",
        " TOP10品牌类目份额占比 (销售额)": sales_vals[1] if len(sales_vals) > 1 else "",
        "品牌总数量": sales_vals[-1] if sales_vals else "",
        "行业类目销量": qty_vals[0] if len(qty_vals) > 0 else "",
    }


def parse_response_payload(html):
    ar_month = json.loads(extract_balanced(html, "ar_month:", "[", "]"))
    vxe_data = json.loads(extract_balanced(html, "vxe_data:", "[", "]"))
    summary = parse_summary_metrics(html)
    return ar_month, vxe_data, summary


def is_incomplete_response(html, top_brand_num):
    try:
        _, tables, summary = parse_response_payload(html)
    except Exception:
        return True, "返回内容无法解析"

    rows = []
    for table in tables:
        rows.extend(table.get("tableData", []))

    if not summary.get("行业类目销售额") or not summary.get("行业类目销量"):
        return True, "行业汇总为空"

    min_expected_rows = 3 if top_brand_num >= 10 else 1
    if len(rows) < min_expected_rows:
        return True, f"品牌数异常偏少：{len(rows)}"

    return False, ""


def wait_for_complete_response_html(get_html, top_brand_num, timeout=30, poll_interval=1):
    deadline = time.time() + timeout
    last_reason = "响应内容为空"

    while time.time() < deadline:
        html = get_html() or ""
        if "vxe_data:" not in html:
            last_reason = "响应中未找到 vxe_data"
        else:
            incomplete, reason = is_incomplete_response(html, top_brand_num)
            if not incomplete:
                return html
            last_reason = reason
        time.sleep(poll_interval)

    raise TimeoutException(f"等待完整响应超时：{last_reason}")


def fmt_pct(value):
    if value in ("", None):
        return ""
    text = str(value).strip()
    return text if text.endswith("%") else f"{text}%"


def build_query_form(url, start_val, end_val, top_brand_num):
    query = parse_qs(urlparse(url).query)
    cid = query.get("cid", [""])[0]
    start_date = (datetime.now() - timedelta(days=14)).strftime("%Y-%m-%d")
    end_date = datetime.now().strftime("%Y-%m-%d")

    base_pairs = [
        ("is_csv", "0"),
        ("s_num", "1"),
        ("s_sales", "0"),
        ("s_hb", "0"),
        ("s_tb", "0"),
        ("s_zb", "0"),
        ("s_zb_20", "0"),
        ("price_start", ""),
        ("price_end", ""),
        ("vxt_num", "0"),
        ("vxt_sales", "1"),
        ("vxt_hb", "0"),
        ("vxt_tb", "0"),
        ("vxt_zb", "0"),
        ("vxt_zb_20", "0"),
        ("vxt_total_fixed", "0"),
    ]

    if CHANNEL == "天猫":
        return [
            ("StatAliNewSearchForm[ecType]", "ali"),
            ("StatAliNewSearchForm[cid]", cid),
            ("StatAliNewSearchForm[alias_bid]", ""),
            ("StatAliNewSearchForm[report_type]", "month"),
            ("StatAliNewSearchForm[startYear]", start_val[:4]),
            ("StatAliNewSearchForm[endYear]", end_val[:4]),
            ("StatAliNewSearchForm[startTime]", start_val),
            ("StatAliNewSearchForm[endTime]", end_val),
            ("StatAliNewSearchForm[startDate]", start_date),
            ("StatAliNewSearchForm[endDate]", end_date),
            ("StatAliNewSearchForm[is_mall][]", "1"),
            ("StatAliNewSearchForm[orderType]", "2"),
            ("StatAliNewSearchForm[top_brand_num]", str(top_brand_num)),
            ("StatAliNewSearchForm[include_other_bid]", "0"),
            *base_pairs,
        ]

    return [
        ("StatNewSearchForm[ecType]", "jd"),
        ("StatNewSearchForm[cid]", cid),
        ("StatNewSearchForm[alias_bid]", ""),
        ("StatNewSearchForm[report_type]", "month"),
        ("StatNewSearchForm[startTime]", start_val),
        ("StatNewSearchForm[endTime]", end_val),
        ("StatNewSearchForm[startDate]", start_date),
        ("StatNewSearchForm[endDate]", end_date),
        ("StatNewSearchForm[shopType][]", "selfcn"),
        ("StatNewSearchForm[shopType][]", "selfglobal"),
        ("StatNewSearchForm[shopType][]", "popcn"),
        ("StatNewSearchForm[shopType][]", "popglobal"),
        ("StatNewSearchForm[orderType]", "2"),
        ("StatNewSearchForm[top_brand_num]", str(top_brand_num)),
        ("is_cross", "0"),
        *base_pairs,
    ]


def fetch_page_html_via_browser_fetch(driver, url, start_val, end_val, top_brand_num=10):
        form_pairs = build_query_form(url, start_val, end_val, top_brand_num)
        script = """
const targetUrl = arguments[0];
const formPairs = arguments[1];
const callback = arguments[arguments.length - 1];

(async () => {
    try {
        const params = new URLSearchParams();
        for (const pair of formPairs) {
            params.append(pair[0], pair[1]);
        }
        const response = await fetch(targetUrl, {
            method: 'POST',
            credentials: 'include',
            redirect: 'follow',
            headers: {
                'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
            },
            body: params.toString()
        });
        const text = await response.text();
        callback(JSON.stringify({ ok: response.ok, status: response.status, text }));
    } catch (error) {
        callback(JSON.stringify({ error: String(error) }));
    }
})();
"""
        result = json.loads(driver.execute_async_script(script, url, form_pairs))
        if result.get("error"):
                raise RuntimeError(f"浏览器 fetch 失败：{result['error']}")
        html = result.get("text", "")
        if not result.get("ok"):
                raise RuntimeError(f"浏览器 fetch 返回状态异常：{result.get('status')}")
        if "vxe_data:" not in html:
                raise RuntimeError("浏览器 fetch 响应中未找到 vxe_data")
        return html


def fetch_page_html_via_navigation(driver, url, start_val, end_val, top_brand_num=10):
    driver.get(url)
    time.sleep(2)

    try:
        if CHANNEL == "天猫":
            cb = WebDriverWait(driver, WAIT).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="f1"]/table/tbody/tr[5]/td[2]/label[2]'))
            )
            inp = driver.find_element(By.XPATH, '//*[@id="f1"]/table/tbody/tr[5]/td[2]/label[2]/input')
            if not inp.is_selected():
                cb.click()
            WebDriverWait(driver, WAIT).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="statalinewsearchform-ordertype"]/label[2]'))
            ).click()
        else:
            checkboxes = WebDriverWait(driver, WAIT).until(
                lambda d: d.find_elements(By.XPATH, '//*[@id="f1"]//input[@type="checkbox"]')
            )
            targets = {"京东国内自营", "京东海外自营", "京东国内POP", "京东海外POP"}
            for inp in checkboxes:
                try:
                    label_text = inp.find_element(By.XPATH, "./parent::label").text.strip().replace(" ", "")
                    if label_text in {t.replace(" ", "") for t in targets} and not inp.is_selected():
                        driver.execute_script("arguments[0].click();", inp)
                except Exception:
                    continue
            WebDriverWait(driver, WAIT).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="statnewsearchform-ordertype"]/label[2]'))
            ).click()

        top = WebDriverWait(driver, WAIT).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="top_brand_num"]'))
        )
        top.clear()
        top.send_keys(str(top_brand_num))

        Select(WebDriverWait(driver, WAIT).until(
            EC.presence_of_element_located((By.ID, "period_start"))
        )).select_by_value(start_val)
        Select(WebDriverWait(driver, WAIT).until(
            EC.presence_of_element_located((By.ID, "period_end"))
        )).select_by_value(end_val)

        WebDriverWait(driver, WAIT).until(
            EC.element_to_be_clickable((By.ID, "submit-button"))
        ).click()
        wait_loading(driver)
        return wait_for_complete_response_html(lambda: driver.page_source, top_brand_num, timeout=30)
    except Exception as exc:
        raise RuntimeError(f"Selenium 页面抓取失败：{exc}") from exc


def fetch_page_html(driver, url, start_val, end_val, top_brand_num=10):
    try:
        html = fetch_page_html_via_browser_fetch(driver, url, start_val, end_val, top_brand_num)
        incomplete, reason = is_incomplete_response(html, top_brand_num)
        if incomplete:
            raise RuntimeError(f"浏览器 fetch 返回疑似不完整数据：{reason}")
        print("     ✓ 已通过浏览器 fetch 直接拿到响应")
        return html
    except Exception as exc:
        print(f"     ℹ 浏览器 fetch 未成功，回退整页导航：{exc}")
        return fetch_page_html_via_navigation(driver, url, start_val, end_val, top_brand_num)


def extract_data_from_html(html, path, ytd_label):
    product = map_product(path)
    month_col = TARGET_MONTH
    ar_month, tables, summary = parse_response_payload(html)
    if TARGET_MONTH not in ar_month:
        print(f"     ⚠ 返回数据中找不到月份 {TARGET_MONTH}，可用月份：{ar_month}")
        return False

    month_idx = ar_month.index(TARGET_MONTH)
    total_idx = len(ar_month) - 1
    record = {
        "渠道": CHANNEL,
        "时间": ytd_label,
        "产品": product,
        "行业类目销售额": summary.get("行业类目销售额", ""),
        "销售额同比": "",
        "行业类目销量": summary.get("行业类目销量", ""),
        "销量同比": "",
        "品牌总数量": summary.get("品牌总数量", ""),
        " TOP10品牌类目份额占比 (销售额)": summary.get(" TOP10品牌类目份额占比 (销售额)", ""),
    }
    new_industry.append(record)
    print(f"     ✓ 行业汇总：销售额={record['行业类目销售额']}  销量={record['行业类目销量']}")

    rows = []
    for table in tables:
        rows.extend(table.get("tableData", []))
    print(f"     📋 月份映射：{ar_month}  品牌数={len(rows)}")

    bosch_found = False
    for row in rows:
        brand = row.get("name", "")
        if not brand:
            continue

        m_qty = str(row.get(f"month_{month_idx}_num", ""))
        m_qty_yoy = fmt_pct(row.get(f"month_{month_idx}_num_tb", ""))
        m_sales = str(row.get(f"month_{month_idx}_sales", ""))
        m_sales_yoy = fmt_pct(row.get(f"month_{month_idx}_sales_tb", ""))
        new_brand.append({
            "渠道": CHANNEL,
            "产品": product,
            "品牌": brand,
            "时间": month_col,
            "销量": m_qty,
            "销量同比": m_qty_yoy,
            "销售额": m_sales,
            "销售额同比": m_sales_yoy,
            "均价": calc_avg(m_sales, m_qty),
        })

        ytd_qty = str(row.get(f"month_{total_idx}_num", ""))
        ytd_qty_yoy = fmt_pct(row.get(f"month_{total_idx}_num_tb", ""))
        ytd_sales = str(row.get(f"month_{total_idx}_sales", ""))
        ytd_sales_yoy = fmt_pct(row.get(f"month_{total_idx}_sales_tb", ""))
        ytd_sales_pct = fmt_pct(row.get(f"month_{total_idx}_sales_zb", ""))
        new_brand.append({
            "渠道": CHANNEL,
            "产品": product,
            "品牌": brand,
            "时间": ytd_label,
            "销量": ytd_qty,
            "销量同比": ytd_qty_yoy,
            "销售额": ytd_sales,
            "销售额同比": ytd_sales_yoy,
            "均价": calc_avg(ytd_sales, ytd_qty),
        })
        new_mkt.append({
            "渠道": CHANNEL,
            "产品": product,
            "品牌": brand,
            "时间": ytd_label,
            "MktShare": ytd_sales_pct,
        })

        if not bosch_found and ("bosch" in brand.lower() or "博世" in brand):
            bosch_found = True
            for rec in reversed(new_industry):
                if rec.get("产品") == product:
                    rec["博世_销售额"] = ytd_sales
                    rec["博世_销售额同比"] = ytd_sales_yoy
                    rec["博世_销量"] = ytd_qty
                    rec["博世_销量同比"] = ytd_qty_yoy
                    rec["博世_市占比 (销售额)"] = ytd_sales_pct
                    break
            print(f"     ✓ 博世：销售额={ytd_sales}  市占比={ytd_sales_pct}")

    if not bosch_found:
        print(f"     ℹ TOP{len(rows)}中未发现博世，博世字段填null")
        for rec in reversed(new_industry):
            if rec.get("产品") == product:
                rec["博世_销售额"] = "null"
                rec["博世_销售额同比"] = "null"
                rec["博世_销量"] = "null"
                rec["博世_销量同比"] = "null"
                rec["博世_市占比 (销售额)"] = "null"
                break

    print(f"     ✓ 品牌：{len(rows)} 个")
    return bosch_found


def parse_prev_year_summary(html, path):
    product = map_product(path)
    summary = parse_summary_metrics(html)
    prev_year_data[product] = {
        "销售额": summary.get("行业类目销售额", ""),
        "销量": summary.get("行业类目销量", ""),
    }
    print(f"     ✓ 往年YTD：销售额={prev_year_data[product]['销售额']}  销量={prev_year_data[product]['销量']}")


# ══════════════════════════════════════════════════════════
# 读取已有数据（只读，不写）
# ══════════════════════════════════════════════════════════
def load_existing_data(src_path):
    if not src_path or not os.path.exists(src_path):
        print("   ℹ 未找到源文件，将新建")
        return [], [], []
    try:
        wb = load_workbook(src_path, data_only=True)
        def read_sheet(name):
            if name not in wb.sheetnames: return []
            ws = wb[name]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            return [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)
                    if any(v is not None for v in r)]
        ei = read_sheet("MktVolumeSummary")
        eb = read_sheet("BrandTrend")
        em = read_sheet("MktShare")
        print(f"   ✓ 历史数据：MktVolumeSummary {len(ei)} 行，BrandTrend {len(eb)} 行，MktShare {len(em)} 行")
        return ei, eb, em
    except Exception as e:
        print(f"   ⚠ 读取失败：{e}"); return [], [], []


# ══════════════════════════════════════════════════════════
# 叶子节点 / 侧边栏
# ══════════════════════════════════════════════════════════
def is_leaf_page(driver):
    try: return any(kw in driver.find_element(By.TAG_NAME, "body").text for kw in LEAF_KEYWORDS)
    except: return False

def has_sub_industry(driver):
    try: return "子行业数据" in driver.find_element(By.TAG_NAME, "body").text
    except: return False

def get_sidebar_links(driver):
    try:
        WebDriverWait(driver, WAIT).until(EC.presence_of_element_located((By.ID, "sidebar_category")))
        time.sleep(1)
        exclude = EXCLUDE_KEYWORDS_JD if CHANNEL == "京东" else EXCLUDE_KEYWORDS_TMALL
        links, seen = [], set()
        for a in driver.find_elements(By.XPATH, '//*[@id="sidebar_category"]//a'):
            try:
                href = a.get_attribute("href") or ""
                name = a.text.strip()
                if not href or not name or href in seen: continue
                if any(kw in name for kw in exclude): continue
                if "cid=" not in href: continue
                if "MBL" in name: continue  # 排除 MBL 分组本身
                # 京东精确排除「机油滤」（不排除「机油滤清器」）
                if CHANNEL == "京东" and name in JD_EXACT_EXCLUDE: continue
                seen.add(href); links.append((name, href))
            except StaleElementReferenceException: continue
        return links
    except TimeoutException: return []

def collect_leaf_urls(driver, url, path_prefix="", depth=0):
    if url in visited_urls: return
    visited_urls.add(url)
    driver.get(url); time.sleep(2)
    indent    = "  " * depth
    sub_links = get_sidebar_links(driver)
    base_url  = url.split("#")[0]
    sub_links = [(n, u) for n, u in sub_links if u.split("#")[0] != base_url]
    print(f"{indent}📂 {path_prefix}（{len(sub_links)} 个子分类）")

    if CHANNEL == "京东":
        for name, child_url in sub_links:
            child_path = f"{path_prefix} > {name}" if path_prefix else name
            leaf_results.append((child_path, child_url))
            print(f"{indent}  ✅ {child_path}")
    elif is_leaf_page(driver):
        leaf_results.append((path_prefix, url))
        print(f"{indent}  ✅ {path_prefix}")
    elif has_sub_industry(driver) or sub_links:
        for name, child_url in sub_links:
            collect_leaf_urls(driver, child_url, f"{path_prefix} > {name}" if path_prefix else name, depth + 1)
            time.sleep(1)
    else:
        print(f"{indent}  ⚠ 跳过：{path_prefix}")


# ══════════════════════════════════════════════════════════
# 解析表头 colid 映射
# ══════════════════════════════════════════════════════════
def parse_col_mapping(driver):
    group_ths = driver.find_elements(By.XPATH,
        "(//table[contains(@class,'vxe-table--header')])[1]//tr[1]/th[contains(@class,'col--group')]")
    sub_ths   = driver.find_elements(By.XPATH,
        "(//table[contains(@class,'vxe-table--header')])[1]//tr[2]/th")
    sub_labels = ["销量","销量占比","销量同比","销售额","销售额占比","销售额同比"]
    COL = {}
    sub_idx = 0
    for g in group_ths:
        if sub_idx >= len(sub_ths): break
        gname   = g.text.strip()
        colspan = int(g.get_attribute("colspan") or 1)
        for j in range(colspan):
            if sub_idx < len(sub_ths):
                COL[f"{gname}_{sub_labels[j%6]}"] = sub_ths[sub_idx].get_attribute("colid")
                sub_idx += 1

    for th in driver.find_elements(By.XPATH,
            "(//table[contains(@class,'vxe-table--header')])[1]//tr[1]/th[@rowspan]"):
        txt, cid = th.text.strip(), th.get_attribute("colid")
        if "排名" in txt and "品牌" not in txt: COL["rank"]  = cid
        elif "品牌名称" in txt:                  COL["brand"] = cid

    print(f"     📋 总计_销售额={COL.get('总计_销售额','')}  总计_销售额同比={COL.get('总计_销售额同比','')}  总计_销售额占比={COL.get('总计_销售额占比','')}")
    return COL

def gcv(driver, row_el, colid):
    if not colid: return ""
    try:
        td = row_el.find_element(By.XPATH, f".//td[@colid='{colid}']")
        return td.text.strip() or td.get_attribute("title") or ""
    except NoSuchElementException: return ""


# ══════════════════════════════════════════════════════════
# 切换表格视图 + 勾选列
# ══════════════════════════════════════════════════════════
def switch_to_table_view(driver):
    for attempt in range(3):
        try:
            a_btn = WebDriverWait(driver, WAIT).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'a[href="#tab_table"]'))
            )
            driver.execute_script("arguments[0].click();", a_btn)
            time.sleep(2)
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".vxe-table--body"))
            )
            print("     ✓ 已切换到表格视图")
            break
        except Exception as e:
            print(f"     ⚠ 表格视图切换失败（第{attempt+1}次）：{e}")
            time.sleep(2)

    try:
        inputs  = driver.find_elements(By.CSS_SELECTOR, ".vxe-checkbox--input")
        TARGETS = {"销量", "销售额", "同比", "占比"}
        for inp in inputs:
            try:
                label = inp.find_element(By.XPATH, "./parent::label")
                if label.text.strip() in TARGETS and not inp.is_selected():
                    driver.execute_script("arguments[0].click();", inp)
            except: continue
        print("     ✓ 列勾选完成，等待表格渲染")
        time.sleep(2)
    except Exception as e:
        print(f"     ⚠ 列勾选失败：{e}")


# ══════════════════════════════════════════════════════════
# 从品牌表格读取行数据
# ══════════════════════════════════════════════════════════
def read_brand_rows(driver):
    fixed_rows = driver.find_elements(By.XPATH,
        "//div[contains(@class,'vxe-table--fixed-left-wrapper')]"
        "//table[contains(@class,'vxe-table--body')]//tr[@rowid]")
    body_rows = driver.find_elements(By.XPATH,
        "//div[contains(@class,'vxe-table--body-wrapper') and contains(@class,'body--wrapper')]"
        "//table[contains(@class,'vxe-table--body')]//tr[@rowid]")
    return fixed_rows, body_rows


# ══════════════════════════════════════════════════════════
# 搜索博世数据（TOP30）
# ══════════════════════════════════════════════════════════
def search_bosch_top30(driver, path, url, ytd_label):
    product = map_product(path)
    year = TARGET_MONTH[:4]
    month = TARGET_MONTH[4:]
    start_val = f"{year}-01"
    end_val = f"{year}-{month}"

    print("     🔍 TOP10未找到博世，改TOP30重新检索...")
    try:
        html = fetch_page_html(driver, url, start_val, end_val, top_brand_num=30)
        ar_month, tables, _ = parse_response_payload(html)
        total_idx = len(ar_month) - 1
        rows = []
        for table in tables:
            rows.extend(table.get("tableData", []))

        bosch = next(
            (row for row in rows if row.get("name") and ("bosch" in row["name"].lower() or "博世" in row["name"])),
            None,
        )
        if not bosch:
            print("     ℹ TOP30中也未发现博世，博世字段填null")
            for rec in reversed(new_industry):
                if rec.get("产品") == product:
                    rec["博世_销售额"] = "null"
                    rec["博世_销售额同比"] = "null"
                    rec["博世_销量"] = "null"
                    rec["博世_销量同比"] = "null"
                    rec["博世_市占比 (销售额)"] = "null"
                    break
            return False

        ytd_sales = str(bosch.get(f"month_{total_idx}_sales", ""))
        ytd_sales_yoy = fmt_pct(bosch.get(f"month_{total_idx}_sales_tb", ""))
        ytd_qty = str(bosch.get(f"month_{total_idx}_num", ""))
        ytd_qty_yoy = fmt_pct(bosch.get(f"month_{total_idx}_num_tb", ""))
        ytd_sales_pct = fmt_pct(bosch.get(f"month_{total_idx}_sales_zb", ""))

        for rec in reversed(new_industry):
            if rec.get("产品") == product:
                rec["博世_销售额"] = ytd_sales
                rec["博世_销售额同比"] = ytd_sales_yoy
                rec["博世_销量"] = ytd_qty
                rec["博世_销量同比"] = ytd_qty_yoy
                rec["博世_市占比 (销售额)"] = ytd_sales_pct
                break
        print(f"     ✓ TOP30找到博世：销售额={ytd_sales}  市占比={ytd_sales_pct}")
        return True
    except Exception as e:
        print(f"     ⚠ TOP30搜索失败：{e}")
        return False


# ══════════════════════════════════════════════════════════
# 提取行业汇总 + 品牌数据
# ══════════════════════════════════════════════════════════
def extract_data(driver, path, url, ytd_label):
    product   = map_product(path)
    month_col = TARGET_MONTH

    try:
        def gt(xpath):
            try: return driver.find_element(By.XPATH, xpath).text.strip()
            except NoSuchElementException: return ""
        BASE = "//*[@id='content']/div[4]/div/div[3]/div[3]/table/tbody"
        record = {
            "渠道": CHANNEL, "时间": ytd_label, "产品": product,
            "行业类目销售额":                gt(f"{BASE}/tr[2]/td[1]"),
            "销售额同比":                    "",
            "行业类目销量":                 gt(f"{BASE}/tr[5]/td[1]"),
            "销量同比":                      "",
            "品牌总数量":                   gt(f"{BASE}/tr[2]/td[5]"),
            " TOP10品牌类目份额占比 (销售额)": gt(f"{BASE}/tr[2]/td[2]"),
        }
        new_industry.append(record)
        print(f"     ✓ 行业汇总：销售额={record['行业类目销售额']}  销量={record['行业类目销量']}")
    except Exception as e:
        print(f"     ⚠ 行业汇总失败：{e}")
        new_industry.append({"渠道": CHANNEL, "时间": ytd_label, "产品": product})

    try:
        js = """
try {
    let tableEl = document.querySelector('.vxe-table--render-default');
    if (!tableEl || !tableEl.__vue__) return JSON.stringify({error: 'no vue instance'});
    let vm = tableEl.__vue__;
    let monthMap = {};
    let totalIdx = null;
    let fixedCols = 3;
    let dataIdx = 0;
    vm.$data.tableGroupColumn.forEach(col => {
        if (col.title && col.title.match(/^\\d{6}$/)) {
            monthMap[col.title] = dataIdx++;
        } else if (col.title === '总计') {
            totalIdx = dataIdx++;
        }
    });
    return JSON.stringify({monthMap, totalIdx, rows: vm.$data.tableData});
} catch(e) { return JSON.stringify({error: e.toString()}); }
"""
        result = json.loads(driver.execute_script(js))
        if "error" in result:
            print(f"     ⚠ Vue 读取失败：{result['error']}")
            return False

        month_map = result["monthMap"]
        total_idx = result["totalIdx"]
        rows      = result["rows"]
        m_idx     = month_map.get(TARGET_MONTH)

        print(f"     📋 月份映射：{month_map}  总计index={total_idx}  品牌数={len(rows)}")

        if m_idx is None:
            print(f"     ⚠ 找不到 {TARGET_MONTH} 对应的月份列")
            return False

        def fmt_pct(v):
            return f"{v}%" if v not in ("", None) else ""

        bosch_found = False
        for row in rows:
            brand = row.get("name", "")
            if not brand: continue

            m_qty       = str(row.get(f"month_{m_idx}_num",      ""))
            m_qty_yoy   = fmt_pct(row.get(f"month_{m_idx}_num_tb",   ""))
            m_sales     = str(row.get(f"month_{m_idx}_sales",    ""))
            m_sales_yoy = fmt_pct(row.get(f"month_{m_idx}_sales_tb", ""))
            new_brand.append({
                "渠道": CHANNEL, "产品": product, "品牌": brand, "时间": month_col,
                "销量": m_qty, "销量同比": m_qty_yoy,
                "销售额": m_sales, "销售额同比": m_sales_yoy,
                "均价": calc_avg(m_sales, m_qty),
            })

            t             = total_idx
            ytd_qty       = str(row.get(f"month_{t}_num",      ""))
            ytd_qty_yoy   = fmt_pct(row.get(f"month_{t}_num_tb",   ""))
            ytd_sales     = str(row.get(f"month_{t}_sales",    ""))
            ytd_sales_yoy = fmt_pct(row.get(f"month_{t}_sales_tb", ""))
            ytd_sales_pct = fmt_pct(row.get(f"month_{t}_sales_zb", ""))
            new_brand.append({
                "渠道": CHANNEL, "产品": product, "品牌": brand, "时间": ytd_label,
                "销量": ytd_qty, "销量同比": ytd_qty_yoy,
                "销售额": ytd_sales, "销售额同比": ytd_sales_yoy,
                "均价": calc_avg(ytd_sales, ytd_qty),
            })

            new_mkt.append({
                "渠道": CHANNEL, "产品": product, "品牌": brand,
                "时间": ytd_label, "MktShare": ytd_sales_pct,
            })

            if not bosch_found and ("bosch" in brand.lower() or "博世" in brand):
                bosch_found = True
                if new_industry:
                    for rec in reversed(new_industry):
                        if rec.get("产品") == product:
                            rec["博世_销售额"]         = ytd_sales
                            rec["博世_销售额同比"]     = ytd_sales_yoy
                            rec["博世_销量"]           = ytd_qty
                            rec["博世_销量同比"]       = ytd_qty_yoy
                            rec["博世_市占比 (销售额)"] = ytd_sales_pct
                            break
                print(f"     ✓ 博世：销售额={ytd_sales}  市占比={ytd_sales_pct}")

        if not bosch_found:
            print(f"     ℹ TOP{len(rows)}中未发现博世，博世字段填null")
            if new_industry:
                for rec in reversed(new_industry):
                    if rec.get("产品") == product:
                        rec["博世_销售额"]         = "null"
                        rec["博世_销售额同比"]     = "null"
                        rec["博世_销量"]           = "null"
                        rec["博世_销量同比"]       = "null"
                        rec["博世_市占比 (销售额)"] = "null"
                        break

        print(f"     ✓ 品牌：{len(rows)} 个")
        return bosch_found

    except Exception as e:
        print(f"     ⚠ 品牌数据失败：{e}")
        traceback.print_exc()
        return False


# ══════════════════════════════════════════════════════════
# 往年同期检索
# ══════════════════════════════════════════════════════════
def scrape_prev_year(driver, path, url):
    product    = map_product(path)
    year       = int(TARGET_MONTH[:4])
    month      = TARGET_MONTH[4:]
    prev_start = f"{year-1}-01"
    prev_end   = f"{year-1}-{month}"

    try:
        print(f"     ✓ 往年时间：{prev_start} ~ {prev_end}")
        html = fetch_page_html(driver, url, prev_start, prev_end, top_brand_num=10)
        parse_prev_year_summary(html, path)
    except Exception as e:
        print(f"     ⚠ 往年数据提取失败：{e}")


# ══════════════════════════════════════════════════════════
# 采集叶子页面
# ══════════════════════════════════════════════════════════
def scrape_leaf_page(driver, path, url):
    print(f"\n  📄 [{CHANNEL}] 采集：{path}")
    year, month = TARGET_MONTH[:4], TARGET_MONTH[4:]
    start_val, end_val = f"{year}-01", f"{year}-{month}"
    ytd_label = get_ytd_label(TARGET_MONTH)
    try:
        print(f"     ✓ 时间：{start_val} ~ {end_val}")
        print("     ✓ 已选择销售额排序")
        if CHANNEL == "天猫":
            print("     ✓ 已勾选天猫全部")
        else:
            print("     ✓ 已勾选京东全部渠道")
        print("     ✓ TOP品牌数已设为10")
        html = fetch_page_html(driver, url, start_val, end_val, top_brand_num=10)
        print("     ✓ 已获取页面源码")
    except Exception as e:
        print(f"     ⚠ 检索失败：{e}")
        return

    bosch_found = extract_data_from_html(html, path, ytd_label)
    if not bosch_found:
        search_bosch_top30(driver, path, url, ytd_label)

    print(f"\n     📅 检索往年同期...")
    scrape_prev_year(driver, path, url)

    product = map_product(path)
    if new_industry and product in prev_year_data:
        for rec in reversed(new_industry):
            if rec.get("产品") == product:
                prev = prev_year_data[product]
                rec["往年YTD"]    = f"{int(year)-1}YTD{month}"
                rec["往年销售额"] = prev["销售额"]
                rec["往年销量"]   = prev["销量"]
                try:
                    cs, ps = parse_num(rec.get("行业类目销售额","0")), parse_num(prev["销售额"])
                    rec["销售额同比"] = f"{round((cs/ps-1)*100,2)}%" if ps > 0 else ""
                except: pass
                try:
                    cq, pq = parse_num(rec.get("行业类目销量","0")), parse_num(prev["销量"])
                    rec["销量同比"] = f"{round((cq/pq-1)*100,2)}%" if pq > 0 else ""
                except: pass
                print(f"     ✓ 往年YTD={rec['往年YTD']}  销售额同比={rec['销售额同比']}  销量同比={rec['销量同比']}")
                break


# ══════════════════════════════════════════════════════════
# 计算 MktShare
# ══════════════════════════════════════════════════════════
def calc_mkt_share(all_industry, all_brand):
    ytd_label = get_ytd_label(TARGET_MONTH)
    year      = TARGET_MONTH[:4]

    industry_sales = defaultdict(dict)
    for row in all_industry:
        t = str(row.get("时间", ""))
        if len(t) == 6 and t.isdigit() and t[:4] == year:
            industry_sales[row.get("产品", "")][t] = parse_num(row.get("行业类目销售额", "0"))

    brand_sales = defaultdict(lambda: defaultdict(dict))
    for row in all_brand:
        t = str(row.get("时间", ""))
        if len(t) == 6 and t.isdigit() and t[:4] == year:
            brand_sales[row.get("产品","")][row.get("品牌","")][t] = parse_num(row.get("销售额","0"))

    result = []
    for product in brand_sales:
        months    = sorted(m for m in industry_sales.get(product,{}) if m <= TARGET_MONTH)
        ind_total = sum(industry_sales[product].get(m, 0.0) for m in months)
        for brand, msales in brand_sales[product].items():
            brand_total = sum(msales.get(m, 0.0) for m in months)
            pct = f"{round(brand_total/ind_total*100,2)}%" if ind_total > 0 else ""
            result.append({"渠道":CHANNEL,"产品":product,"品牌":brand,"时间":ytd_label,"MktShare":pct})
    result.sort(key=lambda x: (x["产品"], x["品牌"]))
    return result


# ══════════════════════════════════════════════════════════
# 保存 Excel
# ══════════════════════════════════════════════════════════
def save_to_excel(output_path, src_path=None):
    print(f"\n💾 保存 → {output_path}")

    if src_path and os.path.exists(src_path):
        shutil.copy2(src_path, output_path)
        print(f"   ✓ 已从源文件复制：{os.path.basename(src_path)}")
    elif not os.path.exists(output_path):
        Workbook().save(output_path)
        print("   ✓ 已新建空文件")

    wb = load_workbook(output_path)

    dfont  = Font(name="Arial", size=10)
    dalign = Alignment(vertical="center")

    def style_row(ws, r):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.font      = dfont
            cell.alignment = dalign

    def append_to_sheet(sheet_name, data, default_headers):
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            for c, h in enumerate(default_headers, 1):
                ws.cell(row=1, column=c, value=h)
        ws      = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        last    = ws.max_row
        for i, row in enumerate(data, 1):
            r = last + i
            for c, h in enumerate(headers, 1):
                if h is None: continue
                ws.cell(row=r, column=c, value=row.get(h, ""))
            style_row(ws, r)
            ws.row_dimensions[r].height = 18

    ytd_label = get_ytd_label(TARGET_MONTH)
    if "BrandTrend" in wb.sheetnames:
        ws_bt = wb["BrandTrend"]
        headers_bt = [ws_bt.cell(1, c).value for c in range(1, ws_bt.max_column + 1)]
        try:
            time_col_idx = headers_bt.index("时间") + 1
            rows_to_delete = []
            for r in range(ws_bt.max_row, 1, -1):
                val = ws_bt.cell(r, time_col_idx).value
                if val and "YTD" in str(val):
                    rows_to_delete.append(r)
            for r in rows_to_delete:
                ws_bt.delete_rows(r)
            if rows_to_delete:
                print(f"   ✓ BrandTrend：已删除旧YTD行 {len(rows_to_delete)} 条")
        except (ValueError, Exception) as e:
            print(f"   ⚠ BrandTrend YTD清理失败：{e}")

    mkt_new = new_mkt

    append_to_sheet("MktVolumeSummary", new_industry, HDR_INDUSTRY)
    append_to_sheet("MktShare",         mkt_new,      HDR_MKT)
    append_to_sheet("BrandTrend",       new_brand,    HDR_BRAND)

    for name in wb.sheetnames:
        if name not in ("MktVolumeSummary", "MktShare", "BrandTrend"):
            print(f"   ✓ 保留原 sheet：{name}")

    print(f"   ✓ MktVolumeSummary：新增 {len(new_industry)} 行")
    print(f"   ✓ MktShare：新增 {len(mkt_new)} 行")
    print(f"   ✓ BrandTrend：新增 {len(new_brand)} 行（含单月+新YTD）")

    wb.save(output_path)
    print(f"   ✓ 已保存 → {output_path}")


# ══════════════════════════════════════════════════════════
# 【新增】爬虫完成后自动调用 build_data.py 生成 data.js
# ══════════════════════════════════════════════════════════
# =====================================================================
# 新增辅助函数（用于清洗数据，替代原 build_data.py）
# =====================================================================
def clean_value(v):
    """把 pandas/numpy 的 NaN、NaT、数值类型转成 JSON 友好的形式。"""
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    if pd.isna(v):
        return ""
    try:
        import numpy as np
        if isinstance(v, (np.integer,)):
            return int(v)
        if isinstance(v, (np.floating,)):
            f = float(v)
            return "" if math.isnan(f) else f
    except ImportError:
        pass
    if isinstance(v, (pd.Timestamp, datetime)):
        return v.strftime("%Y-%m-%d")
    return v

def df_to_records(df: pd.DataFrame) -> list:
    cleaned = df.copy()
    return [
        {col: clean_value(row[col]) for col in cleaned.columns}
        for _, row in cleaned.iterrows()
    ]

# =====================================================================
# 核心报表生成函数（替代原 _auto_build_data_js 逻辑）
# =====================================================================
def _auto_build_data_js(excel_path):
    """
    爬虫保存完 Excel 后，读取数据，清洗，并直接注入 HTML 模版生成单文件报表。
    """
    print(f"{'='*55}")
    print("  🚀 启动自动化报表构建流水线")
    print(f"{'='*55}")

    template_path = Path(__file__).resolve().parent / "report_v3.html"
    if not template_path.exists():
        print(f"❌ 错误：找不到模板文件 {template_path}")
        print("请确保已将 report_template.html 放在该目录下！")
        return

    try:
        import plotly
        plotly_js_path = Path(plotly.__file__).resolve().parent / "package_data" / "plotly.min.js"
        if not plotly_js_path.exists():
            print(f"❌ 错误：找不到本机 Plotly 脚本 {plotly_js_path}")
            print("请先安装 plotly，或检查当前 Python 环境是否完整。")
            return
        plotly_js_content = plotly_js_path.read_text(encoding="utf-8").replace("</script>", "<\\/script>")
    except Exception as exc:
        print(f"❌ 错误：读取本机 Plotly 脚本失败：{exc}")
        print("请确认当前 Python 环境已安装 plotly，用于生成完全离线 HTML。")
        return

    excel_file = Path(excel_path)
    if not excel_file.exists():
        print(f"❌ 错误：刚刚保存的 Excel 文件不存在：{excel_file}")
        return

    try:
        print(f"  > 正在读取：{excel_file.name}")
        df_industry = pd.read_excel(excel_file, sheet_name="MktVolumeSummary")
        df_mkt = pd.read_excel(excel_file, sheet_name="MktShare")
        df_brand = pd.read_excel(excel_file, sheet_name="BrandTrend")

        payload = {
            "industry": df_to_records(df_industry),
            "mkt_share": df_to_records(df_mkt),
            "brand_trend": df_to_records(df_brand),
            "meta": {
                "source_file": excel_file.name,
                "build_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
        }

        json_str = json.dumps(payload, ensure_ascii=False, default=str, separators=(",", ":"))
        json_str = json_str.replace("</", "<\\/").replace("\u2028", "\\u2028").replace("\u2029", "\\u2029")

        print("  > 正在将数据注入 HTML 模版...")
        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()

        final_html = (
            template_content
            .replace('{{PLOTLY_JS}}', plotly_js_content)
            .replace('{{DATA_JSON}}', json_str)
        )

        today = datetime.now().strftime("%Y%m%d")
        output_dir = excel_file.parent
        output_name = output_dir / f"Nint_Scraper_Report_{today}.html"

        with open(output_name, 'w', encoding='utf-8') as f:
            f.write(final_html)

        size_kb = os.path.getsize(output_name) / 1024

        print("\n✅ 全自动流水线执行完毕！")
        print(f"👉 独立报表已生成: {output_name.name} ({size_kb:,.1f} KB)")
        print("💡 以后不再需要 data.js 和 build_data.py，请直接将该 .html 文件上传到 SharePoint 即可。")

    except Exception as e:
        print(f"❌ 生成独立报表时发生错误: {e}")
        import traceback
        traceback.print_exc()

def open_channel_workspace(driver, home_handle, channel):
    data_name = "情报通阿里" if channel == "天猫" else "情报通京东"
    before_handles = list(driver.window_handles)
    driver.switch_to.window(home_handle)
    WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, f'//*[@data-name="{data_name}"]'))
    )
    time.sleep(1)
    driver.execute_script(
        "arguments[0].click();",
        driver.find_element(By.XPATH, f'//*[@data-name="{data_name}"]')
    )
    print(f"   ✓ 已点击{data_name}")
    WebDriverWait(driver, 30).until(lambda d: len(d.window_handles) > len(before_handles))
    new_handle = next(handle for handle in driver.window_handles if handle not in before_handles)
    driver.switch_to.window(new_handle)
    time.sleep(3)
    print(f"   ✓ 已切换到新标签页：{driver.current_url}")
    time.sleep(3)
    return new_handle


def run_channel(driver, home_handle, channel, test_only):
    global CHANNEL, leaf_results
    CHANNEL = channel
    leaf_results = []

    print("="*55); print(f"  Step 4：建立{CHANNEL}采集会话"); print("="*55)
    channel_handle = open_channel_workspace(driver, home_handle, CHANNEL)
    print("   ✓ 后续将直接通过浏览器获取页面源码")
    print(f"✓ 会话已建立\n")

    print("="*55); print(f"  Step 5：加载{CHANNEL}品类列表"); print("="*55)
    url_map = JD_URLS if CHANNEL == "京东" else TMALL_URLS
    for product_name, product_url in url_map.items():
        leaf_results.append((product_name, product_url))
        print(f"   ✓ {product_name}")
    print(f"\n📋 共 {len(leaf_results)} 个品类")

    print("="*55); print(f"  Step 6：采集{CHANNEL}数据（共 {len(leaf_results)} 个）"); print("="*55)
    items_to_scrape = leaf_results[:1] if test_only else leaf_results
    for i, (path, url) in enumerate(items_to_scrape, 1):
        print(f"\n[{i}/{len(items_to_scrape)}]", end=" ")
        scrape_leaf_page(driver, path, url)
        time.sleep(1)

    print("="*55); print(f"  Step 7：检查{CHANNEL}采集结果"); print("="*55)
    print("\n📋 采集结果汇总：")
    for i, (path, url) in enumerate(items_to_scrape, 1):
        product = map_product(path)
        brand_count = len([
            r for r in new_brand
            if r.get("渠道") == CHANNEL and r.get("产品") == product and len(str(r.get("时间", ""))) == 6
        ])
        status = "✅" if brand_count > 0 else "❌ 品牌数=0"
        print(f"  [{i:2d}] {product:<15} 品牌数={brand_count}  {status}")

    while True:
        retry = input(f"\n是否需要重新查询{CHANNEL}某些品类？(Y/N): ").strip().upper()
        if retry == "N":

            break
        if retry == "Y":
            indices_str = input("请输入编号（多个用逗号分隔，如 1,3）：").strip()
            try:
                for idx in [int(x.strip()) for x in indices_str.split(",")]:
                    if 1 <= idx <= len(items_to_scrape):
                        path, url = items_to_scrape[idx - 1]
                        product = map_product(path)
                        new_industry[:] = [
                            r for r in new_industry if not (r.get("渠道") == CHANNEL and r.get("产品") == product)
                        ]
                        new_brand[:] = [
                            r for r in new_brand if not (r.get("渠道") == CHANNEL and r.get("产品") == product)
                        ]
                        new_mkt[:] = [
                            r for r in new_mkt if not (r.get("渠道") == CHANNEL and r.get("产品") == product)
                        ]
                        prev_year_data.pop(product, None)
                        print(f"\n🔄 重新采集 [{idx}] {path}")
                        scrape_leaf_page(driver, path, url)
            except ValueError:
                print("  ⚠ 输入格式错误")

    driver.close()
    driver.switch_to.window(home_handle)


# ══════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════
def main():
    global TARGET_MONTH, CHANNEL

    print("=" * 55)
    print("  nint.com 数据采集工具（增量更新模式）")
    print("=" * 55)
    print(f"配置：HEADLESS={'Y' if HEADLESS else 'N'}  ASSUME_PROXY_READY={'Y' if ASSUME_PROXY_READY else 'N'}")
    if CHROME_USER_DATA_DIR:
        print(f"配置：CHROME_USER_DATA_DIR={CHROME_USER_DATA_DIR}")
    if CHROME_PROFILE_DIR:
        print(f"配置：CHROME_PROFILE_DIR={CHROME_PROFILE_DIR}")

    while True:
        m = input("请输入要采集的月份（格式 YYYYMM，如 202603）：").strip()
        if len(m) == 6 and m.isdigit(): TARGET_MONTH = m; break
        print("格式错误，请重新输入")
    print(f"✓ 目标月份：{TARGET_MONTH}，YTD标签：{get_ytd_label(TARGET_MONTH)}\n")

    print("请输入源文件完整路径（直接拖入文件也可，回车跳过新建）")
    while True:
        src = input("源文件路径：").strip().strip('"').strip("'")
        if src == "": src = None; break
        if os.path.exists(src) and src.endswith(".xlsx"): break
        print("文件不存在或非 .xlsx，请重新输入（或回车新建）")

    if src:
        folder     = os.path.dirname(src)
        name_noext = os.path.splitext(os.path.basename(src))[0]
        new_suffix = f"{TARGET_MONTH[2:4]}YTD{TARGET_MONTH[4:]}"
        new_name   = re.sub(r'\d{2}YTD\d{2}$', new_suffix, name_noext)
        output_path = os.path.join(folder, new_name + ".xlsx")
    else:
        new_suffix  = f"{TARGET_MONTH[2:4]}YTD{TARGET_MONTH[4:]}"
        default_output_dir = Path(__file__).resolve().parent.parent
        output_path = str(default_output_dir / f"RawData_Tmall+JDCombined_{new_suffix}.xlsx")

    print(f"✓ 源文件：{src or '无（新建）'}")
    print(f"✓ 输出文件：{output_path}\n")
    print("✓ 采集渠道：天猫 + 京东\n")

    driver, actual_headless = init_driver_with_fallback()
    try:
        print(f"✓ 浏览器模式：{'无头' if actual_headless else '有头'}\n")
        home_handle = driver.current_window_handle

        print("="*55); print("  Step 1：登录"); print("="*55)
        driver.get(LOGIN_URL); time.sleep(2)
        wait_input(driver, '//*[@id="login-form-app"]/div/div[1]/form/div[1]/input', EMAIL)
        wait_input(driver, '//*[@id="login-form-app"]/div/div[1]/form/div[2]/input', PASSWORD)
        wait_click(driver, '//*[@id="login-form-app"]/div/div[1]/form/button')
        WebDriverWait(driver, WAIT).until(EC.url_changes(LOGIN_URL))
        print("✓ 登录成功\n"); time.sleep(2)

        print("="*55); print("  Step 2：点击情报通入口"); print("="*55)
        wait_click(driver, '/html/body/div/div[2]/div/div[2]/div[1]/div[2]')
        time.sleep(2); print("✓ 弹窗已弹出\n")

        print("="*55); print("  Step 3：切换代理"); print("="*55)
        if ASSUME_PROXY_READY:
            print("✓ 已按配置跳过代理人工确认\n")
        else:
            print("⚠  请现在将代理切换为【国内代理】")
            while input("   已切换代理？(Y): ").strip().upper() != "Y": pass
            print("✓ 代理已确认\n")

        test_only = input("是否只测试第一个品类？(Y=仅测试第一个/N=全部采集): ").strip().upper() == "Y"
        run_channel(driver, home_handle, "天猫", test_only)
        run_channel(driver, home_handle, "京东", test_only)

        # Step 8: 保存 Excel
        save_to_excel(output_path, src)

        # Step 9: 自动生成 data.js（新增）
        _auto_build_data_js(output_path)

        print("\n✅ 完成！文件已保存为", output_path)
        input("按 Enter 键关闭浏览器...")

    except Exception as e:
        print(f"\n❌ 出错：{e}")
        traceback.print_exc()
        driver.save_screenshot("error_screenshot.png")
        if new_industry or new_brand:
            save_to_excel(output_path, src)
            # 出错时不自动生成 data.js，数据可能不完整
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
